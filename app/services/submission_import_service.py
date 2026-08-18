from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any

from fastapi import Request, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.exceptions import DomainError, ErrorCode
from app.models.enums import SubmissionStatus
from app.repositories.rubric_repository import rubric_repository
from app.repositories.student_repository import student_repository
from app.repositories.submission_repository import submission_repository
from app.schemas.submission import (
    SubmissionImportJsonRequest,
    SubmissionImportResponse,
    SubmissionImportRowInput,
    SubmissionImportRowResult,
)
from app.services.assignment_service import assignment_service
from app.services.github_service import (
    GitHubRepositoryCollector,
    GitHubRepositoryRef,
    parse_github_repository_url,
)
from app.services.submission_service import submission_service


@dataclass(frozen=True)
class SubmissionImportCandidate:
    row: int
    student_code: str
    student_name: str
    repository_url: str


@dataclass(frozen=True)
class ValidatedSubmissionImportCandidate:
    candidate: SubmissionImportCandidate
    repository_ref: GitHubRepositoryRef | None
    errors: list[str]


class SubmissionImportService:
    async def parse_request(self, request: Request) -> list[SubmissionImportCandidate]:
        content_type = request.headers.get("content-type", "").split(";")[0].strip().lower()
        body = await request.body()
        if content_type in {"application/json", ""}:
            return self._parse_json(body)
        if content_type in {"text/csv", "application/csv"}:
            return self._parse_csv(body)
        if content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            return self._parse_xlsx(body)
        raise DomainError(
            ErrorCode.VALIDATION_ERROR,
            "Unsupported submission import content type.",
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            details={"content_type": content_type},
        )

    async def import_batch(
        self,
        db: Session,
        assignment_id: int,
        rows: list[SubmissionImportCandidate],
        collector: GitHubRepositoryCollector,
        *,
        dry_run: bool = False,
    ) -> SubmissionImportResponse:
        assignment_service.get_assignment(db, assignment_id)
        locked_rubric = rubric_repository.get_latest_locked_for_assignment(db, assignment_id)
        validated_rows = self._validate_rows(db, assignment_id, rows)
        row_results: list[SubmissionImportRowResult] = []

        for validated in validated_rows:
            candidate = validated.candidate
            if validated.errors:
                row_results.append(
                    SubmissionImportRowResult(
                        row=candidate.row,
                        student_code=candidate.student_code or None,
                        student_name=candidate.student_name or None,
                        repository_url=candidate.repository_url or None,
                        status="INVALID",
                        errors=validated.errors,
                    )
                )
                continue

            repository_url = validated.repository_ref.normalized_url
            if dry_run:
                row_results.append(
                    SubmissionImportRowResult(
                        row=candidate.row,
                        student_code=candidate.student_code,
                        student_name=candidate.student_name,
                        repository_url=repository_url,
                        status="VALID",
                    )
                )
                continue

            student = student_repository.get_by_code(db, candidate.student_code)
            if student is None:
                student = student_repository.create(
                    db,
                    student_code=candidate.student_code,
                    full_name=candidate.student_name,
                )
            submission = submission_repository.create(
                db,
                assignment_id=assignment_id,
                student_id=student.id,
                repository_url=repository_url,
                rubric_id=locked_rubric.id if locked_rubric is not None else None,
                rubric_version_used=(
                    locked_rubric.version if locked_rubric is not None else None
                ),
            )
            try:
                collected = await submission_service.collect_submission(
                    db,
                    submission.id,
                    collector,
                )
            except DomainError as exc:
                row_results.append(
                    SubmissionImportRowResult(
                        row=candidate.row,
                        student_code=candidate.student_code,
                        student_name=candidate.student_name,
                        repository_url=repository_url,
                        status="FAILED",
                        errors=[exc.code.value],
                        submission_id=submission.id,
                    )
                )
                continue

            row_results.append(
                SubmissionImportRowResult(
                    row=candidate.row,
                    student_code=candidate.student_code,
                    student_name=candidate.student_name,
                    repository_url=repository_url,
                    status=self._row_status_for_submission(collected.status),
                    submission_id=collected.id,
                )
            )

        return self._build_response(row_results, dry_run=dry_run)

    def _parse_json(self, body: bytes) -> list[SubmissionImportCandidate]:
        try:
            import_data = SubmissionImportJsonRequest.model_validate_json(body)
        except ValueError:
            try:
                import_data = SubmissionImportJsonRequest(
                    submissions=[
                        SubmissionImportRowInput.model_validate(item)
                        for item in _load_json_list(body)
                    ]
                )
            except ValueError as exc:
                raise DomainError(
                    ErrorCode.VALIDATION_ERROR,
                    "Submission import JSON payload is invalid.",
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                ) from exc
        return [
            SubmissionImportCandidate(
                row=index,
                student_code=(item.student_code or "").strip(),
                student_name=(item.student_name or "").strip(),
                repository_url=(item.repository_url or "").strip(),
            )
            for index, item in enumerate(import_data.submissions, start=1)
        ]

    def _parse_csv(self, body: bytes) -> list[SubmissionImportCandidate]:
        try:
            text = body.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise DomainError(
                ErrorCode.VALIDATION_ERROR,
                "Submission import CSV must be UTF-8 encoded.",
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            ) from exc
        reader = csv.DictReader(StringIO(text))
        return [
            self._candidate_from_mapping(row_index, row)
            for row_index, row in enumerate(reader, start=2)
        ]

    def _parse_xlsx(self, body: bytes) -> list[SubmissionImportCandidate]:
        try:
            workbook = load_workbook(filename=BytesIO(body), read_only=True, data_only=True)
        except Exception as exc:
            raise DomainError(
                ErrorCode.VALIDATION_ERROR,
                "Submission import XLSX payload is invalid.",
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            ) from exc
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        candidates: list[SubmissionImportCandidate] = []
        for row_index, values in enumerate(rows[1:], start=2):
            mapping = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
            }
            candidates.append(self._candidate_from_mapping(row_index, mapping))
        workbook.close()
        return candidates

    def _candidate_from_mapping(
        self,
        row_index: int,
        mapping: dict[str, Any],
    ) -> SubmissionImportCandidate:
        normalized = {
            str(key or "").strip().casefold(): "" if value is None else str(value).strip()
            for key, value in mapping.items()
        }
        return SubmissionImportCandidate(
            row=row_index,
            student_code=normalized.get("student_code", ""),
            student_name=normalized.get("student_name", ""),
            repository_url=normalized.get("repository_url", ""),
        )

    def _validate_rows(
        self,
        db: Session,
        assignment_id: int,
        rows: list[SubmissionImportCandidate],
    ) -> list[ValidatedSubmissionImportCandidate]:
        seen_student_codes: set[str] = set()
        seen_repository_urls: set[str] = set()
        validated_rows: list[ValidatedSubmissionImportCandidate] = []

        for candidate in rows:
            errors: list[str] = []
            repository_ref = None
            normalized_student_code = candidate.student_code.casefold()

            if not candidate.student_code:
                errors.append("MISSING_STUDENT_CODE")
            elif normalized_student_code in seen_student_codes:
                errors.append("DUPLICATE_STUDENT_CODE")
            else:
                seen_student_codes.add(normalized_student_code)

            if not candidate.student_name:
                errors.append("MISSING_STUDENT_NAME")

            if not candidate.repository_url:
                errors.append("MISSING_REPOSITORY_URL")
            else:
                try:
                    repository_ref = parse_github_repository_url(candidate.repository_url)
                except DomainError:
                    errors.append(ErrorCode.INVALID_GITHUB_URL.value)

            if repository_ref is not None:
                if repository_ref.normalized_url.casefold() in seen_repository_urls:
                    errors.append("DUPLICATE_REPOSITORY_URL")
                else:
                    seen_repository_urls.add(repository_ref.normalized_url.casefold())
                if submission_repository.get_for_assignment_repository(
                    db,
                    assignment_id=assignment_id,
                    repository_url=repository_ref.normalized_url,
                ):
                    errors.append("DUPLICATE_REPOSITORY_SUBMISSION")

            if candidate.student_code:
                student = student_repository.get_by_code(db, candidate.student_code)
                if student is not None:
                    existing_name = _normalize_name(student.full_name)
                    imported_name = _normalize_name(candidate.student_name)
                    if existing_name != imported_name:
                        errors.append("STUDENT_NAME_CONFLICT")
                    elif submission_repository.get_for_assignment_student(
                        db,
                        assignment_id=assignment_id,
                        student_id=student.id,
                    ):
                        errors.append("DUPLICATE_SUBMISSION")

            validated_rows.append(
                ValidatedSubmissionImportCandidate(
                    candidate=candidate,
                    repository_ref=repository_ref,
                    errors=errors,
                )
            )
        return validated_rows

    def _row_status_for_submission(self, submission_status: str) -> str:
        if submission_status == SubmissionStatus.INVENTORIED.value:
            return "INVENTORIED"
        if submission_status == SubmissionStatus.FAILED.value:
            return "FAILED"
        return "IMPORTED"

    def _build_response(
        self,
        rows: list[SubmissionImportRowResult],
        *,
        dry_run: bool,
    ) -> SubmissionImportResponse:
        invalid = sum(row.status == "INVALID" for row in rows)
        failed = sum(row.status == "FAILED" for row in rows)
        return SubmissionImportResponse(
            total=len(rows),
            valid=len(rows) - invalid,
            invalid=invalid,
            imported=sum(row.submission_id is not None for row in rows),
            failed=failed,
            dry_run=dry_run,
            rows=rows,
        )


def _normalize_name(value: str) -> str:
    return " ".join(value.casefold().split())


def _load_json_list(body: bytes) -> list[Any]:
    import json

    parsed = json.loads(body)
    if not isinstance(parsed, list):
        raise ValueError("Expected a JSON list.")
    return parsed


submission_import_service = SubmissionImportService()
